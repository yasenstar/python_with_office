from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A2"] = "a2"
ws2 = wb.create_sheet("Sheet2")
ws2["A1"] = "This is sheet2"

wb.security.workbookPassword = "123"
wb.security.lockStructure = True
# wb.security.lockWindow = True

wb.save("./6-11-1.xlsx")