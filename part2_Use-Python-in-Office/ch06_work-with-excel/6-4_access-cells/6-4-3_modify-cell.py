from openpyxl import load_workbook
wb = load_workbook("./test.xlsx")
ws = wb.active

# cell = ws.cell(6,5)
# print(cell.value)

# cell.value = 100
# print(cell.value)

cell2 = ws["F10"]
print(cell2.value)

# cell2.value = 24
# print(cell2.value)

print(ws["F10"].value)