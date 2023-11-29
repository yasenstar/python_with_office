from openpyxl import load_workbook
wb = load_workbook("./test.xlsx")
ws = wb.active

# row_cells = ws[2]
# print(row_cells)

# col_cells = ws["B"]
# print(col_cells)

# row_range_cells = ws[2:5]
# print(row_range_cells)

# col_range_cells = ws["B:D"]
# print(col_range_cells)

# range_cells = ws["b2:f5"]
# print(range_cells)

# cells = ws.iter_rows(min_col = 3, max_col=4, min_row=1, max_row=3)
# for cell in cells:
#     print(cell)

# cells = ws.iter_cols(min_col = 3, max_col=4, min_row=1, max_row=3)
# for cell in cells:
#         print(cell)

cells = ws.iter_cols(min_col = 1, max_col=20, min_row=1, max_row=20, values_only=True)
for cell in cells:
        print(cell)

