from openpyxl import load_workbook
wb = load_workbook("./test.xlsx")
ws = wb.active

cols = ws["S:U"]
col_s = cols[0]
# print(col_a)
col_t = cols[1]
col_u = cols[2]

for i in range(len(col_t)):
    col_u[i].value = (col_s[i].value + col_t[i].value) / 2

print(col_u)

# wb.save("./test_result.xlsx")