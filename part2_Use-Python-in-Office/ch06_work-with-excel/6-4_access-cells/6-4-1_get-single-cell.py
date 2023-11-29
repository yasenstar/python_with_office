from openpyxl import Workbook
wb = Workbook()
print(type(wb))
ws = wb.active
print(type(ws))

cell1 = ws["a6"]
print(type(cell1))

cell2 = ws.cell(1,2)
print(type(cell2))

cell3 = ws.cell(2,5, "hello world")
print(type(cell3))
print(cell3.value)      