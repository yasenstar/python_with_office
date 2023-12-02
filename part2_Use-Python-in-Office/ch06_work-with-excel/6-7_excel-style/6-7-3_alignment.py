from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active

ws.row_dimensions[2].height = 50

ws.column_dimensions["B"].width = 30

ws["A1"] = "Hello"
ws["B2"] = "World World World World World World World World World"

ws["B2"].alignment = Alignment(
    horizontal = 'center',
    vertical = 'center',
    text_rotation = 0,
    wrap_text = False,
    shrink_to_fit = True,
    indent = 0
)

wb.save("./6-7-3.xlsx")