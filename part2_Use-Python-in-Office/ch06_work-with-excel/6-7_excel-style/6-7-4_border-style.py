from openpyxl import Workbook
from openpyxl.styles import Border, Side

wb = Workbook()
ws = wb.active

ws.column_dimensions["B"].width = 50

ws["B2"] = "Hello World"

side = Side(
    style = "thin",
    color = "ff66dd"
)

ws["B2"].border = Border(
    top = side,
    bottom = Side(style="double"),
    left = side,
    right = Side(style="thick", color = "0000FF")
)

wb.save("./6-7-4.xlsx")