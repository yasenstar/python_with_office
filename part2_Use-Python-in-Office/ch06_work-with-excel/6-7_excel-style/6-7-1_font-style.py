from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active

font = Font(
    name = "微软雅黑",
    size = 45,
    color = "0000FF",
    bold = True,
    italic = True,
    strike = True,
    underline = 'doubleAccounting'
)

ws["B2"] = "Hello World"
ws["B2"].font = font

wb.save("./6-7-1.xlsx")