from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, GradientFill

wb = Workbook()
ws = wb.active

ws.column_dimensions["B"].width = 50

ws["B2"] = "Hello World"

side = Side(
    style = "thin",
    color = "ff66dd"
)

ws["B2"].border = Border(
    top = side,
    bottom = Side(style="double"),
    left = side,
    right = Side(style="thick", color = "0000FF")
)

fill = PatternFill(
    patternType = "lightDown",
    fgColor = "F562A4",
    bgColor = "0000FF"
)

ws["B2"].fill = fill

ws["B4"].fill = GradientFill(
    degree = 60,
    stop = ("FFFFFF", "00FF00", "000000")
)

wb.save("./6-7-5.xlsx")