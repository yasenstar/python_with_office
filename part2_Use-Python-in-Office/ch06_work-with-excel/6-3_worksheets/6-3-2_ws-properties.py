from openpyxl import load_workbook
wb = load_workbook("D:\\GitHub\\python_with_office\\part2_Use-Python-in-Office\\ch06_work-with-excel\\6-3_worksheets\\vehicle.xlsx")
ws = wb.active
print(ws.title)

ws.title = "2017年12月份北美洲汽车产量"
print(ws.title)

print(ws.max_row)
print(ws.max_column)
print(ws.dimensions)
print(ws.encoding)

# wb.save("D:\\GitHub\\python_with_office\\part2_Use-Python-in-Office\\ch06_work-with-excel\\6-3_worksheets\\vehicle.xlsx")