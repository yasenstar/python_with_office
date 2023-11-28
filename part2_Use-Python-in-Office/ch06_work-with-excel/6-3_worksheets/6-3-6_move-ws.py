from openpyxl import Workbook

wb = Workbook()

wb.create_sheet("sheet2")
wb.create_sheet("sheet3")
wb.create_sheet("sheet4")
wb.create_sheet("sheet5")

print(wb.sheetnames)

wb.move_sheet("Sheet",1)
print(wb.sheetnames)

wb.move_sheet("sheet4",-3)
print(wb.sheetnames)