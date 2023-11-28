from openpyxl import load_workbook
wb = load_workbook("D:\\GitHub\\python_with_office\\part2_Use-Python-in-Office\\ch06_work-with-excel\\6-3_worksheets\\vehicle.xlsx")

# print(wb.get_sheet_names())

# print(wb.sheetnames)    # recommended method

sheet1 = wb.get_sheet_by_name("2017年12月份北美洲汽车产量")
print(sheet1.title)

sheet2 = wb["2023年2月份韩国汽车（分车型）销量"]    # recommended
print(sheet2.title)

# sheet3 = wb[2]    # not working like this way
# print(sheet3.title)