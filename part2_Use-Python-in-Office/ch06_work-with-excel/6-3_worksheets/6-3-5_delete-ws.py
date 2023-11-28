from openpyxl import Workbook

wb = Workbook()
print(wb.sheetnames)

ws2 = wb.create_sheet("Sheet2")
print(wb.sheetnames)

wb.create_sheet("Sheet3")
wb.create_sheet("Sheet4")
print(wb.sheetnames)

wb.remove(ws2)
print(wb.sheetnames)

del wb["Sheet3"]
print(wb.sheetnames)