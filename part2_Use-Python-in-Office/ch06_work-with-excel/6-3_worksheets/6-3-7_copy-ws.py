from openpyxl import Workbook

wb = Workbook()

print(wb.sheetnames)

ws = wb.active
ws2 = wb.copy_worksheet(ws)

print(wb.sheetnames)

ws2.title = "new sheet"

print(wb.sheetnames)

wb.save("./test.xlsx")