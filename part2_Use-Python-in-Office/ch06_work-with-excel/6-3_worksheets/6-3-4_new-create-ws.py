from openpyxl import load_workbook
wb = load_workbook("D:\\GitHub\\python_with_office\\part2_Use-Python-in-Office\\ch06_work-with-excel\\6-3_worksheets\\vehicle.xlsx")

print(wb.sheetnames)

new_sheet1 = wb.create_sheet("new sheet 1", 0)
print(wb.sheetnames)

new_sheet2 = wb.create_sheet("new sheet 2")
print(wb.sheetnames)

print(new_sheet1.title)
print(new_sheet2.title)

wb.save("D:\\GitHub\\python_with_office\\part2_Use-Python-in-Office\\ch06_work-with-excel\\6-3_worksheets\\vehicle.xlsx")