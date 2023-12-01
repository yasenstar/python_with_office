from openpyxl import Workbook
from openpyxl.formula.translate import Translator

wb = Workbook()
ws = wb.active

ws.append(["Price1", "Price2", "Total Price", "Average Price"])
ws.append([22,63])
ws.append([11,88])
ws.append([15,68])

ws["c2"] = "=SUM(A2,B2)"
ws["d2"] = "=AVERAGE(A2:B2)"

for cell in ws["C3:D4"]:
    translator_sum = Translator(formula="=SUM(A2,B2)", origin="C2")
    translator_avg = Translator(formula="=AVERAGE(A2,B2)", origin="D2")
    
    cell[0].value = translator_sum.translate_formula(cell[0].coordinate)
    cell[1].value = translator_avg.translate_formula(cell[1].coordinate)

wb.save("./6-6-3_2.xlsx")