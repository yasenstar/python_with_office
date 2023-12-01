from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.append(["Price1", "Price2", "Total Price", "Average Price"])
ws.append([22,63])
ws.append([11,88])
ws.append([15,68])

# ws["c2"] = "=SUM(A2,B2)"
# ws["d2"] = "=AVERAGE(A2:B2)"
# ws["d3"] = "=AVERAGE(A3,B3)"

for i in range(2,5):
    ws["c" + str(i)] = "=SUM(A" + str(i) + ",B" + str(i) + ")"
    ws["d" + str(i)] = "=AVERAGE(A" + str(i) + ",B" + str(i) + ")"

wb.save("./6-6-2.xlsx")