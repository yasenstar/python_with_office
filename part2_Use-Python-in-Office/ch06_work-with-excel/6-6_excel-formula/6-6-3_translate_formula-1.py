from openpyxl import Workbook
from openpyxl.formula.translate import Translator

wb = Workbook()
ws = wb.active

ws.append(["Price1", "Price2", "Total Price", "Average Price"])
ws.append([22,63])
ws.append([11,88])
ws.append([15,68])

ws["c2"] = "=SUM(A2,B2)"
ws["d2"] = "=AVERAGE(A2:B2)"

translator = Translator(formula="=SUM(A2,B2)", origin="C2")

ws["c3"] = translator.translate_formula("C3")
ws["c4"] = translator.translate_formula("C4")

wb.save("./6-6-3_1.xlsx")