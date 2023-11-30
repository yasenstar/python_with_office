from openpyxl import Workbook
wb = Workbook()
ws = wb.active

i = 1
for x in range(1,11):
    for y in range(1,21):
        ws.cell(row=x, column=y, value = i)
        i += 1

ws.merge_cells("B2:F6")
ws.merge_cells("I5:K9")

merged_cells = ws.merged_cells
print(type(merged_cells))

for merged_cell_range in merged_cells:
    print("coord:", merged_cell_range.coord)


# wb.save("./test_result-merge-cell-3.xlsx")