from openpyxl import Workbook
wb = Workbook()
ws = wb.active

i = 1
for x in range(1,11):
    for y in range(1,21):
        ws.cell(row=x, column=y, value = i)
        i += 1

ws.merge_cells("B2:F6")
ws.merge_cells(start_row=7,start_column=8, end_row=9, end_column=14)


wb.save("./test_result-merge-cell-3.xlsx")