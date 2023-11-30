from openpyxl import Workbook
wb = Workbook()
ws = wb.active

i = 1
for x in range(1,11):
    for y in range(1,21):
        ws.cell(row=x, column=y, value = i)
        i += 1

print("B2: ", ws["B2"].value)
print("E3: ", ws["E3"].value)
ws.merge_cells("B2:F6")
print("E3: ", ws["E3"].value)
ws.unmerge_cells("B2:F6")
print("E3: ", ws["E3"].value)

wb.save("./test_result-merge-cell.xlsx")