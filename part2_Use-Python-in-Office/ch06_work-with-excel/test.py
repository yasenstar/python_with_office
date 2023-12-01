from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.append(["价格1","价格2","总和","平均值"])
ws.append([22,63])
ws.append([11,88])
ws.append([15,68])
# ws["c2"] = "=SUM(A2,B2)"  # 求和
# ws["d2"] = "=AVERAGE(A2:B2)"  # 求平均值
for i in range(2,5):
    ws["c" + str(i)] = "=SUM(A" + str(i) + ",B" + str(i) + ")"
wb.save("./test.xlsx")